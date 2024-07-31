import tkinter as tk
from tkinter import ttk
from tkinter import filedialog as fd
import tkinter.messagebox as tkmb
import os
import openpyxl as pyxl
from openpyxl import load_workbook
import re

from Window.Frames.ParseFile.Constants import *


class ParseFile(tk.Frame):
    def __init__(self, core, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self.core = core
        self.name_space_label = tk.Label(self, text='Название месторождения*', justify=tk.CENTER)
        self.name_space_entry = tk.Entry(self, justify=tk.CENTER, font='bold')
        self.open_file_button = tk.Button(self, text='Выбрать файл', command=(self._open_file_command))
        self.open_file_entry = tk.Entry(self, width=120)
        self.load_file_button = tk.Button(self, text='Загрузить файл', command=(self._load_file_command))
        self.save_file_button = tk.Button(self, text='Сохранить файл', command=(self._save_file_command))
        self.progressbar = ttk.Progressbar(self, orient='horizontal', length=100, value=100)
        self.table = ttk.Treeview(self, columns=COLUMNS, displaycolumns='#all', show='headings', height=20)
        for i in range(len(COLUMNS)):
            self.table.heading(i, text=COLUMNS[i][0])
            self.table.column(i, width=COLUMNS[i][1], minwidth=COLUMNS[i][1])
        self.size_of_horizont_label = tk.Label(self, text='Размер горизонта', justify=tk.CENTER)
        self.size_of_horizont_entry = tk.Entry(self, justify=tk.CENTER, state=tk.DISABLED)
        self.max_of_horizont_label = tk.Label(self, text='Максимальная высота горизонта', justify=tk.CENTER)
        self.max_of_horizont_entry = tk.Entry(self, justify=tk.CENTER, state=tk.DISABLED)
        self.ore_types_label = tk.Label(self, text='Типы пород')
        self.ore_types_text = tk.Text(self, width=10, wrap=tk.WORD, state=tk.DISABLED)
        self.load_data_button = tk.Button(self, text='Начать работу с данными', command=(self._start_))
        self.precision = -2
        self.precision_var = tk.IntVar(value=self.precision)
        self.precision_scroll = tk.Scale(self, orient=tk.HORIZONTAL, from_=6, to=-10, variable=self.precision_var, command=self._change_precision)
        self.precision_label = tk.Label(self, text='0.01')
        self.recalculate_button = tk.Button(self, text='Отобразить точность', command=(self._recalculate_command))
        self.data = {
            'table': [],
            'name_space': '',
            'horizont_size': 0,
            'max_horizont': 0,
            'ore_types': [],
            'places': []
        }
        self.ore_types = []
        self._pack()
    
    def _pack(self):
        self.columnconfigure(1, weight=1)
        self.rowconfigure(1, weight=1)
        self.name_space_label.grid(         column=1, row=0, columnspan=5, sticky=tk.NSEW)
        self.name_space_entry.grid(         column=1, row=1, columnspan=5, sticky=tk.NSEW)
        self.open_file_button.grid(         column=1, row=2, sticky=tk.NSEW)
        self.open_file_entry.grid(          column=2, row=2, columnspan=2, sticky=tk.NSEW)
        self.load_file_button.grid(         column=4, row=2, sticky=tk.NSEW)
        self.save_file_button.grid(         column=5, row=2, sticky=tk.NSEW)
        self.progressbar.grid(              column=1, row=3, columnspan=5, sticky=tk.NSEW)
        self.size_of_horizont_label.grid(   column=1, row=4, columnspan=2, sticky=tk.NSEW)
        self.size_of_horizont_entry.grid(   column=1, row=5, columnspan=2, sticky=tk.NSEW)
        self.max_of_horizont_label.grid(    column=3, row=4, columnspan=2, sticky=tk.NSEW)
        self.max_of_horizont_entry.grid(    column=3, row=5, columnspan=2, sticky=tk.NSEW)
        self.table.grid(                    column=1, row=6, columnspan=4, sticky=tk.NSEW)
        self.precision_scroll.grid(         column=2, row=7, columnspan=4, sticky=tk.NSEW)
        self.precision_label.grid(          column=1, row=7, columnspan=1, sticky=tk.NSEW)
        self.ore_types_label.grid(          column=5, row=4, rowspan=2, sticky=tk.NSEW)
        self.ore_types_text.grid(           column=5, row=6, rowspan=1, sticky=tk.NSEW)
        self.recalculate_button.grid(       column=1, row=8, columnspan=5, sticky=tk.NSEW)
        self.load_data_button.grid(         column=1, row=10, columnspan=5, sticky=tk.NSEW)
        
    def _open_file_command(self):
        filename = fd.askopenfilename(defaultextension='.xlsx', filetypes=[('All files','*.*'), 
                                                                            ('Excel documents','*.xlsx')])
        self.open_file_entry.delete(0, tk.END)
        self.open_file_entry.insert(0, filename)

    def _load_file_command(self):
        filename = self.open_file_entry.get()
        if not os.path.exists(filename):
            tkmb.showerror('Ошибка чтения файла', 'Файл не найден, укажите настоящий файл')
            return
        try:
            workbook = pyxl.load_workbook(filename)
        except:
            tkmb.showerror('Ошибка чтения файла', 'Проверьте правильность формата файла, не могу прочитать')
            return
        worksheet = workbook[workbook.sheetnames[0]]
        self.data['table'].clear()
        array = []
        last_horizont = 10 ** 6
        last_namespace = ''
        for row in range(2, worksheet.max_row+1):
            array.clear()
            self.data['table'].append([])
            for column in range(1, worksheet.max_column+1):
                cell = worksheet.cell(row, column).value
                # if type(cell) in (int, float):
                #     array.append(round(cell, -self.precision))
                # else:
                #     array.append(cell)
                array.append(cell)
                self.data['table'][row-2].append(cell)
            nameplace = array[0]
            if nameplace not in self.data['places']:
                self.data['places'].append(nameplace)
            namespace = '_'.join(array[0].split('_')[:-1])
            if namespace not in self.data['name_space'].split(' | '):
                if len(self.data['name_space']) < 1:
                    self.data['name_space'] = namespace
                else:
                    self.data['name_space'] += ' | ' + namespace
            horizont = array[1]
            d_horizont = last_horizont - horizont
            if horizont > self.data['max_horizont']:
                self.data['max_horizont'] = horizont
            if last_horizont == 10 ** 6:
                pass
            elif last_namespace != array[0]:
                last_horizont = 10 ** 6
            elif self.data['horizont_size'] == 0:
                self.data['horizont_size'] = d_horizont
            elif not (self.data['horizont_size'] == d_horizont or d_horizont == 0):
                tkmb.showwarning('Внимание', 'Разный размер горизонтов')
            last_horizont = horizont
            last_namespace = array[0]
            ore_type = array[2]
            if ore_type not in self.data['ore_types']:
                self.data['ore_types'].append(ore_type)
            self.__update_states()
            self.table.insert("", row-2, values=array)
            self.progressbar['value'] = int(row * 100 / (worksheet.max_row-1))
            self.update()

    def __update_states(self):
        self.name_space_entry.delete(0, tk.END)
        self.name_space_entry.insert(0, self.data['name_space'])
        self.size_of_horizont_entry['state'] = tk.NORMAL
        self.size_of_horizont_entry.delete(0, tk.END)
        self.size_of_horizont_entry.insert(0, str(self.data['horizont_size']))
        self.size_of_horizont_entry['state'] = tk.DISABLED
        self.max_of_horizont_entry['state'] = tk.NORMAL
        self.max_of_horizont_entry.delete(0, tk.END)
        self.max_of_horizont_entry.insert(0, str(self.data['max_horizont']))
        self.max_of_horizont_entry['state'] = tk.DISABLED
        self.ore_types_text['state'] = tk.NORMAL
        self.ore_types_text.delete(0.0, tk.END)
        self.ore_types_text.insert(0.0, ' '.join(self.data['ore_types']))
        self.ore_types_text['state'] = tk.DISABLED

    def _save_file_command(self):
        pass

    def _start_(self):
        self.core.clean()
        self.core.set(table=self.data['table'], 
                              name_space=self.data['name_space'],
                              ore_types=self.data['ore_types'],
                              places=self.data['places'])
        self.master._start_calculation()

    def _recalculate_command(self):
        self._fill_table()

    def _change_precision(self, precision):
        self.precision = int(precision)
        self.precision_label.config(text=str(10 ** self.precision))

    def _clear_table(self):
        for i in self.table.get_children():
            self.table.delete(i)

    def _fill_table(self):
        self._clear_table()
        for i in range(len(self.data['table'])):
            self.table.insert("", i, values=list(map(
                lambda x: round(x, -self.precision) if type(x) in (int, float) else x, 
                self.data['table'][i]
                )))
            self.update()
