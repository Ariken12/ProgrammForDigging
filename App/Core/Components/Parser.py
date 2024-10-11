import openpyxl as pyxl
import Core.Core as Core


class Parser:
    def __init__(self, core: Core):
        self.core = core

    def __call__(self, filename):
        try:
            workbook = pyxl.load_workbook(filename)
        except:
            yield -1
        worksheet = workbook[workbook.sheetnames[0]]
        self.core.clear()
        components = []
        for i in range(7, worksheet.max_column+1):
            cell = worksheet.cell(1, i).value
            if cell is not None:
                components.append(str(cell))
        self.core['component_types'] = tuple(components)
        array = []
        self.last_horizont = None
        self.last_namespace = ''
        for row in range(2, worksheet.max_row+2):
            array.clear()
            self.core['table'].append([])
            for column in range(1, 6+len(components)+1):
                cell = worksheet.cell(row, column).value
                if (column == 1) and (cell is None):
                    break
                if cell is None:
                    cell = 0
                array.append(cell)
                self.core['table'][row-2].append(cell)
            if len(array) != 6+len(components):
                self.core['table'].pop()
                continue
            if None in array:
                continue
            if (err := self.check_nameplace(array[0]) is not None):
                yield err
            if (err := self.check_horizont(array[1]) is not None):
                yield err
            if (err := self.check_ore(array[2]) is not None):
                yield err
            for i, component in enumerate(self.core['component_types']):
                if (err := self.check_components(component, array[6+i]) is not None):
                    yield err

            yield int(row * 100 / (worksheet.max_row+1))
        while True:
            yield 100
        

    def check_nameplace(self, nameplace):
        self.namespace = nameplace
        if self.namespace not in self.core['namespace'].split(' | '):
            if len(self.core['namespace']) < 1:
                self.core['namespace'] = self.namespace
            else:
                self.core['namespace'] += ' | ' + self.namespace
        if nameplace not in self.core['places']:
            self.core['places'] += (nameplace,)
            self.core['meta']['places'][nameplace] = {
                'V': 0, 
                'M': 0,
                'MIN_H': float('inf'),
                'MAX_H': float('-inf'),
                'STEP_H': -1
            }
        self.core['meta']['places'][nameplace]['V'] += self.core['table'][-1][3]
        self.core['meta']['places'][nameplace]['M'] += self.core['table'][-1][4]

    def check_horizont(self, horizont):
        self.core['meta']['places'][self.namespace]['MIN_H'] = min(self.core['meta']['places'][self.namespace]['MIN_H'], horizont)
        self.core['meta']['places'][self.namespace]['MAX_H'] = max(self.core['meta']['places'][self.namespace]['MAX_H'], horizont)
        if self.last_horizont is None:
            self.last_horizont = horizont
            return None
        d_horizont = self.last_horizont - horizont
        if self.last_namespace != self.namespace or d_horizont == MAX_H - horizont:
            self.last_horizont = horizont
            self.last_namespace = self.namespace
            return None
        self.last_horizont = horizont
        self.last_namespace = self.namespace
        if self.core['meta']['places'][self.namespace]['STEP_H'] == -1 and d_horizont != 0:
            self.core['meta']['places'][self.namespace]['STEP_H'] = d_horizont
        elif not (self.core['meta']['places'][self.namespace]['STEP_H'] == d_horizont or d_horizont == 0):
            return -2

    def check_ore(self, ore):
        if ore not in self.core['ore_types']:
            self.core['ore_types'] += (ore,)
            self.core['meta']['ores'][ore] = {'V': 0, 'M': 0}
        self.core['meta']['ores'][ore]['V'] += self.core['table'][-1][3]
        self.core['meta']['ores'][ore]['M'] += self.core['table'][-1][4]

    def check_components(self, component, sg):
        if component not in self.core['meta']['components']:
            self.core['meta']['components'][component] = 0
        self.core['meta']['components'][component] += self.core['table'][-1][4] * sg 
        